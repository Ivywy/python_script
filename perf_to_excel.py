
import re
import sys
import os
from xml.etree import ElementTree

import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy

from util import common

from bs4 import BeautifulSoup

class ValueNotFoundError(Exception):
    pass

def read_txt(inputpath,app):
    li = list()
    dic={}
    with open(inputpath, "r") as f:
        line = f.readlines()
        for i in line:
            score=re.findall("Score=\d+", i)    # ['Score=505']
            # print("score",score)
            for j in score:
                li.append(re.findall(r"\d+\.?\d*",j))   # [['505']]
                # print("li",li)
    dic[app]=li[0][0]    # 只抽取第一个score
    return dic

def read_html(file_path,app):
    doc = open(file_path, 'r', encoding='utf-8').read()
    soup = BeautifulSoup(doc, "html.parser")
    total=soup.findAll(text=re.compile('.*?Total.*?'))
    # print(total)
    s=""
    dic={}
    for str in total:
        if str.find("Total scores")!=-1:
            s=str.split("Total scores: ")[1]
            break
    dic[app]=float(s)
    return dic

            # for s in re.findall(r'-?\d+\.?\d*', str):
            #     print(s)
def read_xml(file_path,app,item):
    data=dict()
    dom = ElementTree.parse(file_path)
    root = dom.getroot()
    # print(root)
    result = root.findall(".//result/")
    # print(result)
    for element in result:
        # print(element.tag)
        # print("===============")
        # data[element.tag]=element.text
        if element.tag == item:
            data[app] = element.text
            break
    return data
def write_head(fileName,sheetName,row_head,column_head):
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheetName,cell_overwrite_ok=True)

    # 设置字体加粗
    style = xlwt.easyxf('font: bold on')
    # 写入首行
    # for i in range(1,len(row_head)+1):
    #     worksheet.write(0,i,row_head[i-1],style)
    for index, title in enumerate(row_head):
        worksheet.write(0, index+1, title,style)
    # 写入首列
    # for i in range(1,len(column_head)+1):
    #     worksheet.write(i,0,column_head[i-1],style)
    for row in range(len(column_head)):
        worksheet.write(row + 1, 0, column_head[row],style)
    workbook.save(fileName)

def get_position(fileName,sheet_name,row_name,col_name):
    wb = xlrd.open_workbook(fileName,formatting_info=True)
    sheet=wb.sheet_by_name(sheet_name)
    row, col = 0, 0
    row_head=sheet.row_values(0)
    if row_name in row_head:
        col=row_head.index(row_name)
    else:
        raise ValueNotFoundError("%s not found in list"%row_name)
    coloumn_head=sheet.col_values(0)
    # print("column_head=",coloumn_head)
    if col_name in coloumn_head:
        row=coloumn_head.index(col_name)
    else:
        raise ValueNotFoundError("%s not found in list"%col_name)
    return row,col

def write2excel(fileName,sheetName,data):
    wb = xlrd.open_workbook(fileName,formatting_info=True)
    sheetnames =wb.sheet_names()
    # print(sheetnames.index(sheetName))
    workbook = copy(wb=wb)
    worksheet=workbook.get_sheet(sheetnames.index(sheetName))
    for x in sorted(data.keys()):
        for y in sorted(data[x].keys()):
            pos=list(get_position(fileName,sheetName,x,y))
            print(pos)
            # print(data[x][y])
            worksheet.write(pos[0],pos[1],data[x][y])
    workbook.save(fileName)


if __name__ == "__main__":

    data_path="data"
    #
    # # 读取Furmark的score
    # furmark_path=os.path.join(data_path,"FurMark-Scores.txt")
    # print(read_txt(furmark_path,"Furmark"))
    #
    # 读取heaven11的score
    heaven_path=os.path.join(data_path,"heaven11_log.html")
    print(read_html(heaven_path,"Heaven11"))
    #
    # # 读取Firestrike的数据
    # firestrike_path=os.path.join(data_path,"Result-Firestrike.xml")
    # print(read_xml(firestrike_path,"FireStrike","FireStrikeCustomGraphicsScore"))
    #
    # # 读取timespy的数据
    # timespy_path=os.path.join(data_path,"Result-TimespyExtreme.xml")
    # print(read_xml(timespy_path,"TimeSpy","TimeSpyExtremeCustomGraphicsScore"))
    #
    # # 读取3Dmark11的数据
    # file_3Dmark11=os.path.join(data_path,"3Dmark11.xml")
    # print(read_xml(file_3Dmark11,"3Dmark11","graphicsscore"))


    # row_head=["AC + HG","DC + HG","AC + NoHG","DC + NoHG"]
    # column_head=["TimeSpy","Furmark","Heaven11","FireStrike"]
    # write_head("text.xls","sheet1",row_head,column_head)

    # print(find_row_position("text.xls","sheet1","AC + HG","TimeSpy"))

    # data={"AC + HG":{"TimeSpy":1986,"Furmark":505,"Heaven11":1365.35,"FireStrike":3859},"DC + HG":{"TimeSpy":1987,"Furmark":506,"Heaven11":1366.36,"FireStrike":3860},"AC + NoHG":{"TimeSpy":1988,"Furmark":507,"Heaven11":1366.37,"FireStrike":3861},"DC + NoHG":{"TimeSpy":1989,"Furmark":508,"Heaven11":1369.39,"FireStrike":3862}}
    data={"AC + HG":{'Heaven11': 1999.99}}
    write2excel("text.xls","sheet1",data)







